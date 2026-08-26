import openpyxl, os
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Source file
src_file = '3. KAIZEN - BM.xlsx'
wb_src = openpyxl.load_workbook(src_file, data_only=True)
ws_src = wb_src['DATABASE-KAIZEN'] # Exact sheet name in 3. KAIZEN - BM.xlsx

# Dates mapping for Ngày gửi from report files
dates_map = {
    'KZ2025-001': '14/04/2025',
    'KZ2025-002': '14/04/2025',
    'KZ2025-003': '14/04/2025',
    'KZ2025-004': '14/04/2025',
    'KZ2025-005': '21/04/2025',
    'KZ2025-006': '21/05/2025',
    'KZ2025-007': '21/05/2025',
    'KZ2025-008': '21/05/2025',
    'KZ2025-009': '21/05/2025',
    'KZ2025-010': '21/05/2025',
    'KZ2025-011': '21/05/2025',
    'KZ2025-013': '21/05/2025',
    'KZ2025-014': '21/05/2025',
    'KZ2025-015': '21/05/2025',
    'KZ2025-027': '20/08/2025',
}

# Calculation descriptions mapping
calc_map = {
    'KZ2025-001': 'Số tiền tiết kiệm sạn khi sử dụng phương án mới/ 1 vòng đời 1 lò: (50~75 kg) * 37.000đ ≈ 2.300.000đ. Số tiền tiết kiệm được: 2.300.000 đ * 2,5 lò * 6 (mẻ/ngày) * 300 (ngày) / 110 (chu kỳ thay sạn) = 94.090.000đ. Tiền thưởng = 94.090.000đ * 20% = 18.820.000đ',
    'KZ2025-002': 'Cải tiến chi phí đầu tư thấp, mang lại hiệu quả an toàn và ổn định sản xuất. Ban Giám đốc quyết định khen thưởng 2.000.000 đ.',
    'KZ2025-003': 'Tiết kiệm chi phí mua dầu mới và chi phí xử lý dầu thải. Ban Giám đốc quyết định khen thưởng 2.000.000 đ.',
    'KZ2025-004': 'Tần suất sửa chữa máng điện cẩu trục xoay do va chạm năm 2024: 2 lần/năm. Số công sửa chữa 1 lần: 1,5 công. Số tiền mua ray thay thế: 116.000 đ/m * 6m * 2 lần/năm = 1.392.000 đ. Giá trị làm lợi: (1,5 công/lần * 2 lần/năm * 350.000 đ/công + 1.392.000 đ) * 5 cẩu = 12.210.000 đ. Tiền thưởng = 12.210.000 đ * 20% = 2.442.000 đ.',
    'KZ2025-005': 'Số lượng lưỡi cưa sử dụng hàng năm = 32 cái. Theo phương án mới tiết kiệm 16 cái/năm và 6,4 công/năm. Tiền mua lưỡi tiết kiệm: 16 * 1.280.000 = 20.480.000 đ/năm. Tiền công tiết kiệm: 6,4 * 350.000 = 2.240.000 đ. Giá trị làm lợi = (2.240.000 + 20.480.000 - 6.800.000) = 15.920.000 đ. Tiền thưởng = 15.920.000 * 20% = 3.184.000 đ.',
    'KZ2025-006': 'Tiết kiệm thời gian gá lắp và tăng năng suất máy phay cổng. Ban Giám đốc quyết định khen thưởng 3.000.000 đ.',
    'KZ2025-007': 'Tiết kiệm nhân công và tăng sản lượng. Ban Giám đốc quyết định khen thưởng 2.000.000 đ.',
    'KZ2025-008': 'Xét thêm phần lợi ích về tăng năng lực công đoạn, Ban Giám đốc quyết định mức thưởng 5.000.000 đ cho cải tiến này.',
    'KZ2025-009': 'Ban Giám đốc quyết định mức khen thưởng = 2.000.000 đ.',
    'KZ2025-010': 'Ban Giám đốc ra quyết định khen thưởng 1.000.000 đ.',
    'KZ2025-011': 'Áp dụng vào đơn hàng dự kiến 25.000 sp/năm. Ban Giám đốc ra quyết định khen thưởng 5.000.000 đ.',
    'KZ2025-013': 'Dự kiến tổng sản lượng năm 2025 đạt 4.086 sp. Tiền nhân công tiết kiệm: 350.000 * 136 công = 47.670.000đ. Tiết kiệm keo nến: 2.000 * 4.086 = 8.172.000đ. Giá trị làm lợi = 47.670.000 + 8.172.000 - 4.672.000 = 51.170.000đ. Khen thưởng = 51.170.000 * 20% ≈ 10.000.000đ.',
    'KZ2025-014': 'Dự kiến sản lượng hàng TMD năm 2025 đạt 10.362 sp. Tiền nhân công tiết kiệm: 345 công * 350.000đ = 120.890.000đ. Tiết kiệm keo nến: 2.000đ * 2 cây/sp * 10.362 sp = 41.448.000đ. Giá trị làm lợi = 120.890.000 + 41.448.000 - 102.000.000 = 60.338.000đ. Khen thưởng = 60.338.000 * 20% ≈ 12.000.000đ.',
    'KZ2025-015': 'Dự kiến tổng sản lượng năm 2025 đạt 4.086 sp. Tiền nhân công tiết kiệm được khi áp dụng giải pháp mới: 350.000đ * 85 công ≈ 29.794.000đ. Tiền thưởng = 29.794.000 * 20% ≈ 6.000.000đ.',
    'KZ2025-027': 'Giá trị làm lợi được từ cải tiến là 25.000.000đ. Ban Giám đốc ra quyết định khen thưởng 5.000.000đ.',
}

img_dir = 'excel_images'

rows_data = []

# Rows 7 to 21 in 3. KAIZEN - BM.xlsx contain the records
for r in range(7, ws_src.max_row + 1):
    code = ws_src.cell(r, 6).value # Col F (6)
    if not code:
        continue
    code = str(code).strip()
    
    idea = ws_src.cell(r, 3).value # Col C (3)
    name = ws_src.cell(r, 4).value # Col D (4)
    unit = ws_src.cell(r, 5).value # Col E (5)
    status = ws_src.cell(r, 7).value # Col G (7)
    status_quo = ws_src.cell(r, 12).value # Col L (12)
    solution = ws_src.cell(r, 13).value # Col M (13)
    resources = ws_src.cell(r, 16).value # Col P (16)
    benefits = ws_src.cell(r, 17).value # Col Q (17)
    notes_eval = ws_src.cell(r, 18).value # Col R (18)
    expansion = ws_src.cell(r, 19).value # Col S (19)
    reward_status = ws_src.cell(r, 20).value # Col T (20)
    value_vnd = ws_src.cell(r, 21).value # Col U (21)
    reward_vnd = ws_src.cell(r, 22).value # Col V (22)
    notes_extra = ws_src.cell(r, 23).value # Col W (23)
    
    # Image file paths
    img_before_path = f"{img_dir}/{code}_before.png"
    img_after_path = f"{img_dir}/{code}_after.png"
    
    # Check if images exist, construct hyperlink formula
    if os.path.exists(img_before_path):
        link_before = f'=HYPERLINK("{img_before_path}", "🔗 Xem hình trước ({code})")'
    else:
        link_before = 'Không có hình'
        
    if os.path.exists(img_after_path):
        link_after = f'=HYPERLINK("{img_after_path}", "🔗 Xem hình sau ({code})")'
    else:
        link_after = 'Không có hình'
        
    calc_desc = calc_map.get(code, '')
    date_sent = dates_map.get(code, '')
    dept_impl = 'Ban Cải tiến'
    
    # Combine notes if appropriate
    combined_notes = str(notes_eval or '').strip()
    if notes_extra and str(notes_extra).strip():
        combined_notes = (combined_notes + '\n' + str(notes_extra).strip()).strip()

    rows_data.append({
        'code': code,
        'link_before': link_before,
        'link_after': link_after,
        'name': name,
        'unit': unit,
        'idea': idea,
        'status_quo': status_quo,
        'solution': solution,
        'benefits': benefits,
        'resources': resources,
        'calc_desc': calc_desc,
        'expansion': expansion,
        'status': status,
        'reward_status': reward_status,
        'date_sent': date_sent,
        'dept_impl': dept_impl,
        'notes': combined_notes,
        'value_vnd': value_vnd if isinstance(value_vnd, (int, float)) else None,
        'reward_vnd': reward_vnd if isinstance(reward_vnd, (int, float)) else None,
        'reward_date': ''
    })

# Create new workbook
wb_new = openpyxl.Workbook()
ws_new = wb_new.active
ws_new.title = 'Tổng hợp Kaizen BM'
ws_new.views.sheetView[0].showGridLines = True

headers = [
    'Mã ý tưởng', 'Hình trước', 'Hình sau', 'Họ và tên', 'Đơn vị', 
    'Ý tưởng', 'Thực trạng', 'Giải pháp', 'Lợi ích mang lại', 'Nguồn lực sử dụng', 
    'Mô tả cách tính', 'Cơ hội nhân rộng phát triển', 'Trạng thái', 'Tình trạng khen thưởng', 'Ngày gửi', 
    'Phòng ban triển khai', 'Ghi chú', 'Giá trị làm lợi (VND)', 'Tiền thưởng (VND)', 'Ngày duyệt khen thưởng'
]

col_widths = {
    'A': 16, # Mã ý tưởng
    'B': 28, # Hình trước (Link)
    'C': 28, # Hình sau (Link)
    'D': 20, # Họ và tên
    'E': 22, # Đơn vị
    'F': 32, # Ý tưởng
    'G': 42, # Thực trạng
    'H': 42, # Giải pháp
    'I': 42, # Lợi ích mang lại
    'J': 28, # Nguồn lực sử dụng
    'K': 45, # Mô tả cách tính
    'L': 35, # Cơ hội nhân rộng
    'M': 16, # Trạng thái
    'N': 20, # Tình trạng khen thưởng
    'O': 14, # Ngày gửi
    'P': 20, # Phòng ban triển khai
    'Q': 35, # Ghi chú
    'R': 22, # Giá trị làm lợi
    'S': 20, # Tiền thưởng
    'T': 22  # Ngày duyệt
}

font_header = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
fill_header = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
align_header = Alignment(horizontal='center', vertical='center', wrap_text=True)

font_data = Font(name='Segoe UI', size=10, color='000000')
font_code = Font(name='Segoe UI', size=10, bold=True, color='1F4E78')
font_link = Font(name='Segoe UI', size=10, color='0563C1', underline='single')
font_num = Font(name='Segoe UI', size=10, bold=True, color='006100')

fill_even = PatternFill(start_color='F4F7FA', end_color='F4F7FA', fill_type='solid')
fill_odd = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

align_top_left = Alignment(horizontal='left', vertical='top', wrap_text=True)
align_top_center = Alignment(horizontal='center', vertical='top', wrap_text=True)
align_top_right = Alignment(horizontal='right', vertical='top', wrap_text=True)

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

# Header Row
ws_new.row_dimensions[1].height = 32
for c_idx, h_text in enumerate(headers, 1):
    cell = ws_new.cell(row=1, column=c_idx, value=h_text)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_header
    cell.border = thin_border

# Data Rows
for r_idx, rec in enumerate(rows_data, start=2):
    ws_new.row_dimensions[r_idx].height = 45
    fill_row = fill_even if r_idx % 2 == 0 else fill_odd
    
    row_values = [
        rec['code'],
        rec['link_before'],
        rec['link_after'],
        rec['name'],
        rec['unit'],
        rec['idea'],
        rec['status_quo'],
        rec['solution'],
        rec['benefits'],
        rec['resources'],
        rec['calc_desc'],
        rec['expansion'],
        rec['status'],
        rec['reward_status'],
        rec['date_sent'],
        rec['dept_impl'],
        rec['notes'],
        rec['value_vnd'],
        rec['reward_vnd'],
        rec['reward_date']
    ]
    
    for c_idx, val in enumerate(row_values, 1):
        cell = ws_new.cell(row=r_idx, column=c_idx, value=val)
        cell.font = font_data
        cell.fill = fill_row
        cell.border = thin_border
        
        if c_idx == 1: # Code
            cell.font = font_code
            cell.alignment = align_top_center
        elif c_idx in [2, 3]: # Links for images
            cell.font = font_link
            cell.alignment = align_top_center
        elif c_idx in [13, 14, 15, 16, 20]: # Dates & Statuses
            cell.alignment = align_top_center
        elif c_idx in [18, 19]: # Numbers
            cell.font = font_num
            cell.alignment = align_top_right
            if val is not None:
                cell.number_format = '#,##0'
        else: # Text
            cell.alignment = align_top_left

# Apply column widths
for col_letter, width in col_widths.items():
    ws_new.column_dimensions[col_letter].width = width

output_filename = 'Tong_hop_Kaizen_BM_Link.xlsx'
wb_new.save(output_filename)
print(f'Successfully generated {output_filename} with {len(rows_data)} records!')
