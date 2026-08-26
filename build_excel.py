import fitz, openpyxl, os, glob, re
from PIL import Image as PILImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ensure directory for cropped images
img_dir = 'excel_images'
os.makedirs(img_dir, exist_ok=True)

codes = ['KZ2025-001', 'KZ2025-002', 'KZ2025-003', 'KZ2025-004', 'KZ2025-005', 
         'KZ2025-006', 'KZ2025-007', 'KZ2025-008', 'KZ2025-009', 'KZ2025-010', 
         'KZ2025-011', 'KZ2025-013', 'KZ2025-014', 'KZ2025-015', 'KZ2025-027']

for code in codes:
    pdf_f = f'{code}.pdf'
    if not os.path.exists(pdf_f):
        continue
    doc = fitz.open(pdf_f)
    page = doc[0]
    rect = page.rect
    
    if code == 'KZ2025-001':
        clip_bef = fitz.Rect(45, 350, 415, 530)
        clip_aft = fitz.Rect(415, 350, 785, 530)
    elif code in ['KZ2025-004', 'KZ2025-005']:
        clip_bef = fitz.Rect(40, 270, 390, 520)
        clip_aft = fitz.Rect(395, 270, 750, 520)
    elif code in ['KZ2025-013', 'KZ2025-014', 'KZ2025-015', 'KZ2025-027']:
        clip_bef = fitz.Rect(30, 280, 305, 430)
        clip_aft = fitz.Rect(310, 280, 580, 430)
    elif code == 'KZ2025-003':
        clip_bef = fitz.Rect(45, 300, 305, 460)
        clip_aft = fitz.Rect(310, 300, 570, 460)
    else:
        clip_bef = fitz.Rect(30, 310, 295, 530)
        clip_aft = fitz.Rect(300, 310, 570, 530)
        
    pix_bef = page.get_pixmap(dpi=200, clip=clip_bef)
    pix_aft = page.get_pixmap(dpi=200, clip=clip_aft)
    
    path_bef = f'{img_dir}/{code}_before.png'
    path_aft = f'{img_dir}/{code}_after.png'
    pix_bef.save(path_bef)
    pix_aft.save(path_aft)

records = [
    {
        'code': 'KZ2025-001',
        'name': 'Nguyễn Bá Thà',
        'unit': 'Phân xưởng Đúc',
        'idea': 'Phương pháp đầm lò điện cảm ứng mới',
        'status_quo': 'Đầm lớp sàn lò phẳng không tạo góc nghiêng dốc về lỗ ra gang (lỗ tháo nước tính). Khi ra gang cần dùng dưỡng hình chữ V (tạo hình từ sạn đầm lò mới) kê vào miệng lỗ tháo, dẫn tới mất thời gian, gây mất an toàn lao động và giảm tuổi thọ lò. Vị trí đáy lò châm lửa gang ra nhiều nhất là nguyên nhân chính gây thủng lò và giảm tuổi thọ lò.',
        'solution': 'Đầm sàn nghiêng dốc về phía lỗ tháo gang. Tận dụng từ trường của các vòng đồng phía dưới đáy lò nâng cao nhiệt độ khu vực đáy lò.',
        'benefits': 'Phương pháp đầm lò mới tăng năng suất 200 kg/mẻ => 21,5 tấn/ vòng đời 1 lò. Giảm thời gian nấu luyện do tận dụng được từ trường của những vòng đồng phía đáy. Tiết kiệm 50~75 kg sạn/1 vòng đời 1 lò. Tận dụng dưỡng sẵn có. Giữ nguyên tuổi thọ lò. Giảm lượng điện tiêu thụ.',
        'resources': 'Không',
        'calc_desc': 'Số tiền tiết kiệm sạn khi sử dụng phương án mới/ 1 vòng đời 1 lò: (50~75 kg) * 37.000đ ≈ 2.300.000đ. Số tiền tiết kiệm được: 2.300.000 đ * 2,5 lò * 6 (mẻ/ngày) * 300 (ngày) / 110 (chu kỳ thay sạn) = 94.090.000đ. Tiền thưởng = 94.090.000đ * 20% = 18.820.000đ',
        'expansion': 'Có cơ hội nhân rộng ra các lò khác',
        'status': 'Đã hoàn thành',
        'reward_status': 'Đã khen thưởng',
        'date_sent': '14/04/2025',
        'dept_impl': 'Ban Cải tiến',
        'notes': 'Các hiệu quả mang lại bao gồm: Năng suất, Giảm giá thành sản phẩm. Có cơ hội nhân rộng ra các lò khác.',
        'value_vnd': 94090000,
        'reward_vnd': 18820000,
        'reward_date': ''
    },
    {
        'code': 'KZ2025-002',
        'name': 'Ngô Đăng Luận',
        'unit': 'Xưởng gia công',
        'idea': 'Cải tiến lại hệ thống máng chia nước cẩu xoắn',
        'status_quo': 'Đường ống nước mềm làm nguội cụm lô hàn đường ống xoắn thường xuyên bị xoắn quẩy, tắc nước, rách ống, nước phun ra sàn gây mất an toàn và dừng máy sửa chữa.',
        'solution': 'Lắp hệ thống chia nước bằng máng trượt và vòng quay đồng trục giúp ống nước tự di chuyển theo hành trình của cẩu xoắn không bị vặn xoắn.',
        'benefits': 'Đảm bảo lưu thông nước làm mát liên tục, không bị tắc hay rò rỉ nước, đảm bảo an toàn lao động và tăng độ bền đường ống.',
        'resources': 'Tận dụng vật tư thiết bị cũ sẵn có và nhân công bảo dưỡng.',
        'calc_desc': 'Cải tiến chi phí đầu tư thấp, mang lại hiệu quả an toàn và ổn định sản xuất. Ban Giám đốc quyết định khen thưởng 2.000.000 đ.',
        'expansion': 'Có thể nhân rộng cho các cẩu xoắn và thiết bị có đường ống làm mát di động tương tự.',
        'status': 'Đã hoàn thành',
        'reward_status': 'Đã khen thưởng',
        'date_sent': '14/04/2025',
        'dept_impl': 'Ban Cải tiến',
        'notes': 'Tăng tính an toàn và giảm thời gian dừng máy sửa chữa.',
        'value_vnd': None,
        'reward_vnd': 2000000,
        'reward_date': ''
    },
    {
        'code': 'KZ2025-003',
        'name': 'Phùng Gia Cường',
        'unit': 'Phòng Cải tiến',
        'idea': 'Tái sử dụng dầu gọt kim loại làm mát cho máy tiện CNC',
        'status_quo': 'Dầu làm mát máy CNC sau một thời gian sử dụng bị lẫn tạp chất, cặn bẩn phải bỏ đi gây lãng phí chi phí mua dầu mới và chi phí xử lý chất thải.',
        'solution': 'Lắp đặt hệ thống lọc tuần hoàn tận dụng lại dung dịch dầu gọt kim loại, tách cặn bẩn và tạp chất.',
        'benefits': 'Tiết kiệm chi phí mua dầu làm mát mới, giảm lượng chất thải nguy hại ra môi trường, kéo dài tuổi thọ dụng cụ cắt.',
        'resources': 'Vật tư bộ lọc và nhân công lắp đặt.',
        'calc_desc': 'Tiết kiệm chi phí mua dầu mới và chi phí xử lý dầu thải. Ban Giám đốc quyết định khen thưởng 2.000.000 đ.',
        'expansion': 'Áp dụng cho toàn bộ các máy tiện, máy phay CNC trong xưởng gia công.',
        'status': 'Đã hoàn thành',
        'reward_status': 'Đã khen thưởng',
        'date_sent': '14/04/2025',
        'dept_impl': 'Ban Cải tiến',
        'notes': 'Bảo vệ môi trường và giảm chi phí sản xuất.',
        'value_vnd': None,
        'reward_vnd': 2000000,
        'reward_date': ''
    },
    {
        'code': 'KZ2025-004',
        'name': 'Hưng',
        'unit': 'Phòng thiết bị',
        'idea': 'Thanh chắn máng điện cẩu',
        'status_quo': 'Quá trình làm việc cẩu trục xoay, máng điện va chạm với cẩu trục. Gây hỏng hóc máng điện cẩu trục, dừng cẩu trục để sửa chữa. Tần suất sửa chữa: 2 lần/năm/cẩu xoay, cẩu chữ A (3 cẩu xoay Đúc 1, 2 cẩu chữ A Khuôn nền).',
        'solution': 'Lắp đặt thanh chắn bảo vệ cẩu tránh va chạm trực tiếp giữa cẩu trục và máng điện cẩu trục xoay. Loại bỏ tình trạng hỏng hóc máng điện khi vận hành cẩu trục quay từ nguyên nhân va chạm.',
        'benefits': 'Triệt để loại bỏ hỏng hóc máng điện do nguyên nhân va chạm. Đảm bảo an toàn điện khi vận hành cẩu trục. Giảm tần suất, khối lượng công việc phát sinh cho đội sửa chữa.',
        'resources': 'Phân xưởng lên ý tưởng thiết kế, hoàn thiện che máng điện.',
        'calc_desc': 'Tần suất sửa chữa máng điện cẩu trục xoay do va chạm năm 2024: 2 lần/năm. Số công sửa chữa 1 lần: 1,5 công. Số tiền mua ray thay thế: 116.000 đ/m * 6m * 2 lần/năm = 1.392.000 đ. Giá trị làm lợi: (1,5 công/lần * 2 lần/năm * 350.000 đ/công + 1.392.000 đ) * 5 cẩu = 12.210.000 đ. Tiền thưởng = 12.210.000 đ * 20% = 2.442.000 đ.',
        'expansion': 'Có nhiều cơ hội nhân rộng ra các cẩu trục.',
        'status': 'Đã hoàn thành',
        'reward_status': 'Đã khen thưởng',
        'date_sent': '14/04/2025',
        'dept_impl': 'Ban Cải tiến',
        'notes': 'Cải tiến chi phí đầu tư thấp nhưng hiệu quả mang lại cao: An toàn, Tăng năng suất.',
        'value_vnd': 12210000,
        'reward_vnd': 2442000,
        'reward_date': ''
    },
    {
        'code': 'KZ2025-005',
        'name': 'Hoàng Ngọc Hà',
        'unit': 'Xưởng gia công',
        'idea': 'Máy mài răng cưa lưỡi cưa vòng',
        'status_quo': 'Số lần tái sử dụng lưỡi cưa 2-3 lần. Mài từng răng riêng lẻ, thời gian mài 1 răng: 10-15s/lưỡi, 1 người mài tay. Có rủi ro mất an toàn lao động, mài nhầm răng, răng sau mài không đều, thiếu thẩm mỹ, hỏng lưỡi cưa. Tần suất thay lưỡi mới: 2,5 lưỡi/tháng. Số lượng lưỡi cưa sử dụng hàng năm = 32 cái.',
        'solution': 'Trang bị máy mài răng cưa tự động. Số lần tái sử dụng lưỡi cưa tăng lên 4-5 lần. Mài từng răng 1-1,5s/lưỡi, không cần công nhân mài tay. An toàn lao động, mài đúng răng, răng đều, thẩm mỹ. Tần suất thay lưỡi mới giảm còn 1-1,5 lưỡi/tháng (16 cái/năm).',
        'benefits': 'Việc mài răng cưa an toàn hơn, không còn nguy cơ trượt tay gây đứt tay. Các răng mài đều nâng cao tính thẩm mỹ, tăng niên hạn sử dụng lưỡi cưa. Giảm chi phí mua lưỡi cưa mới.',
        'resources': 'Quản đốc phân xưởng phối hợp phòng cải tiến nghiên cứu chọn mua máy (Chi phí 6.800.000 đ).',
        'calc_desc': 'Số lượng lưỡi cưa sử dụng hàng năm = 32 cái. Theo phương án mới tiết kiệm 16 cái/năm và 6,4 công/năm. Tiền mua lưỡi tiết kiệm: 16 * 1.280.000 = 20.480.000 đ/năm. Tiền công tiết kiệm: 6,4 * 350.000 = 2.240.000 đ. Giá trị làm lợi = (2.240.000 + 20.480.000 - 6.800.000) = 15.920.000 đ. Tiền thưởng = 15.920.000 * 20% = 3.184.000 đ.',
        'expansion': 'Có nhiều cơ hội nhân rộng ra các phân xưởng, các công cụ có cùng tính chất.',
        'status': 'Đã hoàn thành',
        'reward_status': 'Đã khen thưởng',
        'date_sent': '21/04/2025',
        'dept_impl': 'Ban Cải tiến',
        'notes': 'Cải tiến chi phí đầu tư thấp nhưng hiệu quả mang lại cao: Năng suất, Giảm giá thành, An toàn.',
        'value_vnd': 15920000,
        'reward_vnd': 3184000,
        'reward_date': ''
    },
    {
        'code': 'KZ2025-006',
        'name': 'Hoàng Ngọc Hà',
        'unit': 'Xưởng gia công',
        'idea': 'Cải tiến bàn gá cắt phay cổng',
        'status_quo': 'Gá lắp sản phẩm trên máy phay cổng mất nhiều thời gian căn chỉnh, gá kẹp không chắc chắn dễ làm lệch kích thước gia công.',
        'solution': 'Thiết kế bàn gá chuẩn hóa với các chốt định vị nhanh và đồ gá kẹp định hình.',
        'benefits': 'Rút ngắn thời gian gá lắp sản phẩm từ 45 phút xuống 15 phút, tăng độ chính xác gia công, đảm bảo an toàn.',
        'resources': 'Vật tư gá kẹp và nhân công gia công đồ gá.',
        'calc_desc': 'Tiết kiệm thời gian gá lắp và tăng năng suất máy phay cổng. Ban Giám đốc quyết định khen thưởng 3.000.000 đ.',
        'expansion': 'Áp dụng cho các dòng máy phay, máy gọt lớn trong phân xưởng.',
        'status': 'Đã hoàn thành',
        'reward_status': 'Đã khen thưởng',
        'date_sent': '21/05/2025',
        'dept_impl': 'Ban Cải tiến',
        'notes': 'Tăng năng suất gia công máy phay cổng.',
        'value_vnd': None,
        'reward_vnd': 3000000,
        'reward_date': ''
    },
    {
        'code': 'KZ2025-007',
        'name': 'Nguyễn Văn Duy',
        'unit': 'PX NL',
        'idea': 'Cải tiến phương pháp đột chân mút nan giường',
        'status_quo': 'Thao tác đột chân mút thủ công mất nhiều công sức, năng suất thấp và đường đột không đồng đều.',
        'solution': 'Chế tạo khuôn đột gá trên máy ép thủy lực/khí nén giúp thao tác đột nhanh và chính xác.',
        'benefits': 'Tăng năng suất đột gấp 2 lần, đường đột sắc nét, giảm mệt mỏi cho công nhân.',
        'resources': 'Vật tư làm khuôn đột và công thợ.',
        'calc_desc': 'Tiết kiệm nhân công và tăng sản lượng. Ban Giám đốc quyết định khen thưởng 2.000.000 đ.',
        'expansion': 'Áp dụng cho các mã sản phẩm nan giường tương tự.',
        'status': 'Đã hoàn thành',
        'reward_status': 'Đã khen thưởng',
        'date_sent': '21/05/2025',
        'dept_impl': 'Ban Cải tiến',
        'notes': 'Cải thiện điều kiện làm việc và tăng năng suất.',
        'value_vnd': None,
        'reward_vnd': 2000000,
        'reward_date': ''
    },
    {
        'code': 'KZ2025-008',
        'name': 'Vũ Văn Nam',
        'unit': 'PX Đúc 1',
        'idea': 'Thùng gom sản phẩm đúc hỏng',
        'status_quo': 'Sản phẩm đúc hỏng vứt rải rác trên sàn xưởng gây mất vệ sinh, cản trở di chuyển, gây nhầm lẫn mã hàng và lãng phí diện tích đốt dế.',
        'solution': 'Thiết kế và bố trí các thùng gom sản phẩm hỏng có phân loại rõ ràng tại từng khu vực đúc.',
        'benefits': 'Giữ sàn xưởng sạch sẽ, gọn gàng, tránh nhầm lẫn phế phẩm với sản phẩm đạt, giải phóng diện tích lưu trữ. Khách hàng đánh giá tốt.',
        'resources': 'Vật tư hàn thùng gom.',
        'calc_desc': 'Xét thêm phần lợi ích về tăng năng lực công đoạn, Ban Giám đốc quyết định mức thưởng 5.000.000 đ cho cải tiến này.',
        'expansion': 'Nhân rộng ra toàn bộ các phân xưởng trong công ty.',
        'status': 'Đã hoàn thành',
        'reward_status': 'Đã khen thưởng',
        'date_sent': '21/05/2025',
        'dept_impl': 'Ban Cải tiến',
        'notes': 'Cải thiện 5S và quản lý trực quan.',
        'value_vnd': None,
        'reward_vnd': 5000000,
        'reward_date': ''
    },
    {
        'code': 'KZ2025-009',
        'name': 'Phạm Duy Quỳnh',
        'unit': 'Phụ trách hàng CR (Hoàn thiện)',
        'idea': 'Tool nối dài đầu máy mài',
        'status_quo': 'Sản phẩm TRS 6030-500 có hộc sản phẩm sâu, đầu kẹp máy mài chai không thể mài bằng dụng cụ có sẵn tại công ty. Thiết bị sản xuất không mài trực tiếp cận mài lỗ sâu của sản phẩm, có thể phải thuê ngoài hoặc mua thêm dụng cụ.',
        'solution': 'Chế tạo tool nối dài cho máy mài chai của phân xưởng từ đầu máy mài chai cũ của các máy đã hỏng không thể sửa chữa.',
        'benefits': 'Giải quyết triệt để việc mài chi tiết chưa thể mài với dụng cụ sẵn có. Giảm chi phí gia công do không phải thuê ngoài hoặc mua dụng cụ mới. Đạt yêu cầu chất lượng khách hàng mong đợi.',
        'resources': 'Đầu máy mài chai từ các máy hỏng + 15 phút công thợ.',
        'calc_desc': 'Ban Giám đốc quyết định mức khen thưởng = 2.000.000 đ.',
        'expansion': 'Có nhiều cơ hội nhân rộng ra các sản phẩm khác có lỗ sâu.',
        'status': 'Đã hoàn thành',
        'reward_status': 'Đã khen thưởng',
        'date_sent': '21/05/2025',
        'dept_impl': 'Ban Cải tiến',
        'notes': 'Tận dụng phế liệu chế tạo dụng cụ hữu ích.',
        'value_vnd': None,
        'reward_vnd': 2000000,
        'reward_date': ''
    },
    {
        'code': 'KZ2025-010',
        'name': 'Bùi Định Hùng',
        'unit': 'PX Hoàn thiện',
        'idea': 'Hàn cắt nối cánh van',
        'status_quo': 'Sản phẩm cánh van sau đúc xuất hiện nhiều lỗi không thể sửa chữa thông thường. Phải đúc mới sản phẩm để đảm bảo chất lượng, làm chậm tiến độ giao hàng, lãng phí vật tư và chi phí sản xuất.',
        'solution': 'Cắt ghép 2 cánh van bị hỏng, hàn nối, hàn bổ sung từ 2 sản phẩm hỏng thành 1 sản phẩm đạt chất lượng yêu cầu của khách hàng.',
        'benefits': 'Đảm bảo tiến độ giao hàng đúng hạn, giảm chi phí sản xuất cho đơn hàng do phát sinh hàng lỗi, giải quyết vật liệu khó tái chế, giải phóng khu vực tồn kho sản phẩm hỏng.',
        'resources': '1 công thợ + Que hàn.',
        'calc_desc': 'Ban Giám đốc ra quyết định khen thưởng 1.000.000 đ.',
        'expansion': 'Ý tưởng giúp giải quyết giao hàng đúng hạn, giảm chi phí khi quy trình đúc chưa đảm bảo.',
        'status': 'Đã hoàn thành',
        'reward_status': 'Đã khen thưởng',
        'date_sent': '21/05/2025',
        'dept_impl': 'Ban Cải tiến',
        'notes': 'Cứu đơn hàng gấp và giảm lãng phí phế phẩm.',
        'value_vnd': None,
        'reward_vnd': 1000000,
        'reward_date': ''
    },
    {
        'code': 'KZ2025-011',
        'name': 'Nguyễn Văn Duy',
        'unit': 'PX NL',
        'idea': 'Phương pháp cắt chân đậu nan giường Đức Tài',
        'status_quo': 'Công nhân phải xếp sản phẩm ra nền, chặt chân đậu bằng máy chặt cầm tay Ø230, mài vết chặt còn thừa, xếp lên pallet. Thao tác cúi người gây mệt mỏi, rủi ro bệnh nghề nghiệp, năng suất thấp (~350 sp/ca), vết cắt không đều, nguy cơ mất an toàn cao do đá mài văng.',
        'solution': 'Sử dụng máy cắt bàn đá D350 thay máy mài cầm tay D230. Chế tạo đồ gá để định vị sản phẩm khi cắt. Đầu chặt rơi xuống máng trượt vào thùng chứa tự động.',
        'benefits': 'Tăng năng suất từ 350 lên 500 sp/công (tăng 43%). Vết cắt phẳng đẹp 100% đạt chất lượng. Loại bỏ công đoạn mài phẳng chân đậu và thu gom. Thao tác đứng giảm mệt mỏi, hướng tia cắt cố định đảm bảo an toàn lao động.',
        'resources': 'Tận dụng vật liệu tái chế hàn cắt tạo hình đường dốc + 1h công thợ.',
        'calc_desc': 'Áp dụng vào đơn hàng dự kiến 25.000 sp/năm. Ban Giám đốc ra quyết định khen thưởng 5.000.000 đ.',
        'expansion': 'Có cơ hội nhân rộng đối với các sản phẩm kích thước nhỏ có mặt cắt chân đậu thoáng.',
        'status': 'Đã hoàn thành',
        'reward_status': 'Đã khen thưởng',
        'date_sent': '21/05/2025',
        'dept_impl': 'Ban Cải tiến',
        'notes': 'Chi phí đầu tư thấp nhưng hiệu quả cao về năng suất và an toàn.',
        'value_vnd': None,
        'reward_vnd': 5000000,
        'reward_date': ''
    },
    {
        'code': 'KZ2025-013',
        'name': 'Phùng Đình Chiến',
        'unit': 'PX Mẫu xốp',
        'idea': 'Thay đổi khuôn cánh xoắn Ishizue',
        'status_quo': 'Sản phẩm Ishizue do thay đổi công nghệ có thêm yêu cầu hàn vành, mẫu trắng sau khi đúc xong phải qua phòng cắt để cắt ngắn sau đó dán viền để đáp ứng yêu cầu hàn. Tốn nhân công phát sinh thêm 2 công đoạn, năng suất 1,5h/sp (~5 sp/ngày/người).',
        'solution': 'Chỉnh sửa khuôn đúc mẫu trắng (tạo vành gờ 5mm*2mm ở mẫu luôn) để sản phẩm đúc xong đạt ngay yêu cầu về hình dáng. Rút ngắn thời gian thao tác 13 phút/sp, năng suất đạt 1,25h/sp (~6 sp/ngày/người).',
        'benefits': 'Cắt giảm 2 công đoạn ~13 phút/sp + chi phí vật tư tiêu hao (keo nến), giảm ~3 phút thời gian xử lý lỗi chất lượng ở công đoạn mông hàng. Độ bám sơn tốt hơn.',
        'resources': 'Nhôm tấm 2mm (972.000đ) + Công cắt dây 4h*10 khuôn (3.000.000đ) + Công lắp vành nhôm (700.000đ). Tổng chi phí sửa khuôn: 4.672.000đ.',
        'calc_desc': 'Dự kiến tổng sản lượng năm 2025 đạt 4.086 sp. Tiền nhân công tiết kiệm: 350.000 * 136 công = 47.670.000đ. Tiết kiệm keo nến: 2.000 * 4.086 = 8.172.000đ. Giá trị làm lợi = 47.670.000 + 8.172.000 - 4.672.000 = 51.170.000đ. Khen thưởng = 51.170.000 * 20% ≈ 10.000.000đ.',
        'expansion': 'Đã áp dụng trên 10 loại sản phẩm Ishizue (165DW500, 190DW500, 216DW650, 216DW550, 267DW700, 267DW800, 318DW800, 318DW650, 355DW750, 355DW900).',
        'status': 'Đã hoàn thành',
        'reward_status': 'Đã khen thưởng',
        'date_sent': '21/05/2025',
        'dept_impl': 'Ban Cải tiến',
        'notes': 'Tăng năng suất và nâng cao chất lượng sản phẩm Ishizue.',
        'value_vnd': 51170000,
        'reward_vnd': 10000000,
        'reward_date': ''
    },
    {
        'code': 'KZ2025-014',
        'name': 'Phùng Đình Chiến',
        'unit': 'PX Mẫu xốp',
        'idea': 'Thay đổi khuôn cánh xoắn TMD',
        'status_quo': 'Mẫu trắng được đúc thành 2 nửa, qua công đoạn gắn sẽ ghép thành sản phẩm hoàn chỉnh. Năng suất đạt 20~25 sp/ca, sản phẩm dễ bị lỗi cong, lồi lõm, sơn không dính ở vị trí bôi keo dẫn đến phát sinh thời gian sửa hàng.',
        'solution': 'Làm lại khuôn để định hình sản phẩm đúng yêu cầu khách hàng, vị trí ghép nối không làm ảnh hưởng đến hình dáng của sản phẩm. Giảm 1 nhân công ngồi gắn, giảm 2 thanh keo nến/sp, năng suất tăng lên 150 sp/ngày.',
        'benefits': 'Tăng năng suất lao động từ 20-25 sp lên 150 sp/ngày, loại bỏ hết các lỗi chất lượng liên quan đến hình dáng kích thước sản phẩm cũ, giảm thời gian phát sinh cho công đoạn sau.',
        'resources': 'Chi phí sửa khuôn (6 bộ) = 102.000.000đ.',
        'calc_desc': 'Dự kiến sản lượng hàng TMD năm 2025 đạt 10.362 sp. Tiền nhân công tiết kiệm: 345 công * 350.000đ = 120.890.000đ. Tiết kiệm keo nến: 2.000đ * 2 cây/sp * 10.362 sp = 41.448.000đ. Giá trị làm lợi = 120.890.000 + 41.448.000 - 102.000.000 = 60.338.000đ. Khen thưởng = 60.338.000 * 20% ≈ 12.000.000đ.',
        'expansion': 'Đã áp dụng trên 4 bộ khuôn cho các sản phẩm: 1012196, 1012197R, 1021059-03, 1021060-03.',
        'status': 'Đã hoàn thành',
        'reward_status': 'Đã khen thưởng',
        'date_sent': '21/05/2025',
        'dept_impl': 'Ban Cải tiến',
        'notes': 'Đột phá về năng suất đúc mẫu xốp cánh xoắn TMD.',
        'value_vnd': 60338000,
        'reward_vnd': 12000000,
        'reward_date': ''
    },
    {
        'code': 'KZ2025-015',
        'name': 'Lê Văn Khương',
        'unit': 'Phòng Thiết bị',
        'idea': 'Sửa đổi chương trình hàn để tăng chiều dày đắp',
        'status_quo': 'Hàn 2 lớp (lặp lại chương trình hàn 2 lần) để hoàn thiện 1 sản phẩm cánh xoắn Ishizue, mất nhiều thời gian vận hành máy.',
        'solution': 'Sửa chương trình hàn, hàn 1 lần hoàn thành, điều chỉnh tốc độ đi chậm hơn để đạt được chiều dày đắp mong muốn ngay trong 1 lần hàn.',
        'benefits': 'Tăng năng lực sản xuất lên 1,5 lần (sản phẩm 457A95 tăng từ 8 sp/nửa ca lên 12 sp/nửa ca).',
        'resources': 'Không (chỉnh sửa phần mềm chương trình hàn).',
        'calc_desc': 'Dự kiến tổng sản lượng năm 2025 đạt 4.086 sp. Tiền nhân công tiết kiệm được khi áp dụng giải pháp mới: 350.000đ * 85 công ≈ 29.794.000đ. Tiền thưởng = 29.794.000 * 20% ≈ 6.000.000đ.',
        'expansion': 'Có thể áp dụng được trên tất cả các sản phẩm cánh xoắn Ishizue.',
        'status': 'Đã hoàn thành',
        'reward_status': 'Đã khen thưởng',
        'date_sent': '21/05/2025',
        'dept_impl': 'Ban Cải tiến',
        'notes': 'Cải tiến phần mềm điều khiển không tốn chi phí nhưng mang lại hiệu quả cao.',
        'value_vnd': 29794000,
        'reward_vnd': 6000000,
        'reward_date': ''
    },
    {
        'code': 'KZ2025-027',
        'name': 'Phạm Văn Nhị',
        'unit': 'PX Đúc 1',
        'idea': 'Đổi ống rót to cho các dòng cánh xoắn to (từ D40 sang D60)',
        'status_quo': 'Sử dụng ống rót D40 cho các dòng cánh xoắn 350DW900, 406DW1000, 457DW950, 457DW1150. Sản phẩm bị lỗi sập khuôn 6,75% (11 sp sập / 163 sp đúc).',
        'solution': 'Tăng ống rót lên D60, tăng chiều sâu đậu bù và góc nghiêng của đậu bù.',
        'benefits': 'Giảm tỷ lệ lỗi hàng hỏng do sập khuôn về 0% (đạt 180/180 sp đúc).',
        'resources': 'Không.',
        'calc_desc': 'Giá trị làm lợi được từ cải tiến là 25.000.000đ. Ban Giám đốc ra quyết định khen thưởng 5.000.000đ.',
        'expansion': 'Xem xét áp dụng giải pháp này trên các sản phẩm có lỗi sập khuôn khác.',
        'status': 'Đã hoàn thành',
        'reward_status': 'Đã khen thưởng',
        'date_sent': '20/08/2025',
        'dept_impl': 'P. Cải tiến',
        'notes': 'Triệt hạ hoàn toàn lỗi sập khuôn cho cánh xoắn lớn.',
        'value_vnd': 25000000,
        'reward_vnd': 5000000,
        'reward_date': ''
    }
]

headers = [
    'Mã ý tưởng', 'Hình trước', 'Hình sau', 'Họ và tên', 'Đơn vị', 
    'Ý tưởng', 'Thực trạng', 'Giải pháp', 'Lợi ích mang lại', 'Nguồn lực sử dụng', 
    'Mô tả cách tính', 'Cơ hội nhân rộng phát triển', 'Trạng thái', 'Tình trạng khen thưởng', 'Ngày gửi', 
    'Phòng ban triển khai', 'Ghi chú', 'Giá trị làm lợi (VND)', 'Tiền thưởng (VND)', 'Ngày duyệt khen thưởng'
]

col_widths = {
    'A': 16, # Mã ý tưởng
    'B': 32, # Hình trước
    'C': 32, # Hình sau
    'D': 20, # Họ và tên
    'E': 22, # Đơn vị
    'F': 32, # Ý tưởng
    'G': 42, # Thực trạng
    'H': 42, # Giải pháp
    'I': 42, # Lợi ích mang lại
    'J': 28, # Nguồn lực sử dụng
    'K': 45, # Mô tả cách tính
    'L': 35, # Cơ hội nhân rộng
    'M': 16, # Trạng thái
    'N': 20, # Tình trạng khen thưởng
    'O': 14, # Ngày gửi
    'P': 20, # Phòng ban triển khai
    'Q': 35, # Ghi chú
    'R': 22, # Giá trị làm lợi
    'S': 20, # Tiền thưởng
    'T': 22  # Ngày duyệt
}

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Tổng hợp ý tưởng cải tiến'
ws.views.sheetView[0].showGridLines = True

# Styling definitions
font_header = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
fill_header = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid') # Navy Blue
align_header = Alignment(horizontal='center', vertical='center', wrap_text=True)

font_data = Font(name='Segoe UI', size=10, color='000000')
font_code = Font(name='Segoe UI', size=10, bold=True, color='1F4E78')
font_num = Font(name='Segoe UI', size=10, bold=True, color='006100')

fill_even = PatternFill(start_color='F4F7FA', end_color='F4F7FA', fill_type='solid')
fill_odd = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

align_top_left = Alignment(horizontal='left', vertical='top', wrap_text=True)
align_top_center = Alignment(horizontal='center', vertical='top', wrap_text=True)
align_top_right = Alignment(horizontal='right', vertical='top', wrap_text=True)

thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

# Write Header Row
ws.row_dimensions[1].height = 32
for c_idx, h_text in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c_idx, value=h_text)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_header
    cell.border = thin_border

# Write Data Rows
for r_idx, rec in enumerate(records, start=2):
    ws.row_dimensions[r_idx].height = 130 # Accommodate images
    fill_row = fill_even if r_idx % 2 == 0 else fill_odd
    
    row_values = [
        rec['code'],
        '', # Hình trước
        '', # Hình sau
        rec['name'],
        rec['unit'],
        rec['idea'],
        rec['status_quo'],
        rec['solution'],
        rec['benefits'],
        rec['resources'],
        rec['calc_desc'],
        rec['expansion'],
        rec['status'],
        rec['reward_status'],
        rec['date_sent'],
        rec['dept_impl'],
        rec['notes'],
        rec['value_vnd'],
        rec['reward_vnd'],
        rec['reward_date']
    ]
    
    for c_idx, val in enumerate(row_values, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        cell.font = font_data
        cell.fill = fill_row
        cell.border = thin_border
        
        # Alignment & Formatting per column type
        if c_idx == 1: # Code
            cell.font = font_code
            cell.alignment = align_top_center
        elif c_idx in [2, 3]: # Images
            cell.alignment = align_top_center
        elif c_idx in [13, 14, 15, 16, 20]: # Dates, Statuses, Depts
            cell.alignment = align_top_center
        elif c_idx in [18, 19]: # Numbers
            cell.font = font_num
            cell.alignment = align_top_right
            if val is not None:
                cell.number_format = '#,##0'
        else: # Text fields
            cell.alignment = align_top_left

    # Insert Before Image (Col B / Col 2)
    path_bef = f"{img_dir}/{rec['code']}_before.png"
    if os.path.exists(path_bef):
        img_b = openpyxl.drawing.image.Image(path_bef)
        img_b.width = 200
        img_b.height = 145
        cell_ref = f'B{r_idx}'
        ws.add_image(img_b, cell_ref)
        
    # Insert After Image (Col C / Col 3)
    path_aft = f"{img_dir}/{rec['code']}_after.png"
    if os.path.exists(path_aft):
        img_a = openpyxl.drawing.image.Image(path_aft)
        img_a.width = 200
        img_a.height = 145
        cell_ref = f'C{r_idx}'
        ws.add_image(img_a, cell_ref)

# Apply column widths
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# Save file
output_filename = 'Tong_hop_bao_cao_cai_tien_VICO.xlsx'
wb.save(output_filename)
print(f'Successfully generated {output_filename} with {len(records)} records!')
